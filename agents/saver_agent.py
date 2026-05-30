"""
SAVER AGENT (Agent 4)
Saves business data to:
  - CSV  format  (lightweight, universal)
  - XLSX format  (professional Excel with formatting)
  - Daily files  (one file per day per category)
"""

import os
import sys
import csv
import pandas as pd
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# Fix import path on Windows
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from config.settings import OUTPUT_DIR, BUSINESS_FIELDS


class SaverAgent:

    HEADERS = [
        "Business Name", "Phone Number", "Email", "Website",
        "Rating", "Review Count", "Address", "City",
        "Category", "Source", "Scraped Date"
    ]

    FIELD_KEYS = [
        "business_name", "phone_number", "email", "website",
        "rating", "review_count", "address", "city",
        "category", "source", "scraped_date"
    ]

    def __init__(self, business_type="businesses", city="unknown"):
        self.business_type = business_type.replace(" ", "_").lower()
        self.city          = city.replace(" ", "_").lower()
        self.today         = str(date.today())
        self.base_name     = f"{self.business_type}_{self.city}_{self.today}"

        os.makedirs(OUTPUT_DIR, exist_ok=True)

        self.csv_path  = os.path.join(OUTPUT_DIR, f"{self.base_name}.csv")
        self.xlsx_path = os.path.join(OUTPUT_DIR, f"{self.base_name}.xlsx")

        self._init_csv()
        self._init_xlsx()

    # ------------------------------------------------------------------ CSV
    def _init_csv(self):
        """Create CSV file with headers if it doesn't exist"""
        if not os.path.exists(self.csv_path):
            with open(self.csv_path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(self.HEADERS)
            print(f"📄 CSV  created: {self.csv_path}")

    def save_to_csv(self, business: dict) -> bool:
        try:
            row = [business.get(k, "") for k in self.FIELD_KEYS]
            with open(self.csv_path, "a", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(row)
            return True
        except Exception as e:
            print(f"❌ CSV save error: {e}")
            return False

    # ----------------------------------------------------------------- XLSX
    def _init_xlsx(self):
        """Create XLSX file with styled headers if it doesn't exist"""
        if not os.path.exists(self.xlsx_path):
            wb = Workbook()
            ws = wb.active
            ws.title = "Businesses"

            # Header style
            header_fill   = PatternFill("solid", fgColor="1F4E79")
            header_font   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
            header_align  = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border   = Border(
                left   = Side(style="thin", color="CCCCCC"),
                right  = Side(style="thin", color="CCCCCC"),
                bottom = Side(style="thin", color="CCCCCC"),
            )

            col_widths = [30, 18, 28, 35, 8, 12, 40, 18, 20, 15, 14]

            for col_idx, (header, width) in enumerate(zip(self.HEADERS, col_widths), 1):
                cell               = ws.cell(row=1, column=col_idx, value=header)
                cell.fill          = header_fill
                cell.font          = header_font
                cell.alignment     = header_align
                cell.border        = thin_border
                ws.column_dimensions[get_column_letter(col_idx)].width = width

            ws.row_dimensions[1].height = 30
            ws.freeze_panes = "A2"

            # Summary sheet
            ws2         = wb.create_sheet("Summary")
            ws2["A1"]   = "GMB Hunter — Daily Summary"
            ws2["A1"].font = Font(name="Arial", bold=True, size=14, color="1F4E79")
            ws2["A3"]   = "Date:"
            ws2["B3"]   = self.today
            ws2["A4"]   = "Category:"
            ws2["B4"]   = self.business_type
            ws2["A5"]   = "City:"
            ws2["B5"]   = self.city
            ws2["A7"]   = "Total Records:"
            ws2["B7"]   = '=COUNTA(Businesses!A2:A10000)'
            ws2["A7"].font = Font(bold=True)

            for cell in ["A3","A4","A5","A7"]:
                ws2[cell].font = Font(name="Arial", bold=True)

            wb.save(self.xlsx_path)
            print(f"📊 XLSX created: {self.xlsx_path}")

    def save_to_xlsx(self, business: dict) -> bool:
        """Append one business record to the XLSX file"""
        try:
            os.makedirs(OUTPUT_DIR, exist_ok=True)
            if not os.path.exists(self.xlsx_path):
                self._init_xlsx()
            wb = load_workbook(self.xlsx_path)
            ws = wb["Businesses"]

            next_row   = ws.max_row + 1
            row_values = [business.get(k, "") for k in self.FIELD_KEYS]

            # Alternate row colors for readability
            row_fill = PatternFill(
                "solid",
                fgColor="EBF3FB" if next_row % 2 == 0 else "FFFFFF"
            )
            data_font  = Font(name="Arial", size=10)
            data_align = Alignment(vertical="center", wrap_text=False)

            for col_idx, value in enumerate(row_values, 1):
                cell           = ws.cell(row=next_row, column=col_idx, value=value)
                cell.fill      = row_fill
                cell.font      = data_font
                cell.alignment = data_align

            ws.row_dimensions[next_row].height = 18
            wb.save(self.xlsx_path)
            return True
        except Exception as e:
            print(f"❌ XLSX save error: {e}")
            return False

    # --------------------------------------------------------- SAVE BOTH
    def save(self, business: dict) -> bool:
        """Save to both CSV and XLSX at once"""
        csv_ok  = self.save_to_csv(business)
        xlsx_ok = self.save_to_xlsx(business)
        return csv_ok and xlsx_ok

    def get_output_paths(self) -> dict:
        return {
            "csv":  self.csv_path,
            "xlsx": self.xlsx_path
        }


if __name__ == "__main__":
    # Quick test
    saver = SaverAgent(business_type="restaurants", city="New York")

    test_business = {
        "business_name": "Test Restaurant",
        "phone_number":  "+1-555-123-4567",
        "email":         "test@restaurant.com",
        "website":       "https://testrestaurant.com",
        "rating":        4.5,
        "review_count":  128,
        "address":       "123 Main St",
        "city":          "New York",
        "category":      "restaurants",
        "source":        "test",
        "scraped_date":  str(date.today()),
    }

    result = saver.save(test_business)
    print(f"✅ Test save result: {result}")
    print(f"📁 Files: {saver.get_output_paths()}")